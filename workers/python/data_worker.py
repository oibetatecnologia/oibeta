import sys
import json
import csv
import os

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

def get_encoding_and_separator(file_path):
    # Simple heuristic
    encodings = ['utf-8', 'latin-1', 'iso-8859-1']
    separators = [',', ';', '\t', '|']
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc) as f:
                first_line = f.readline()
                if first_line:
                    best_sep = ','
                    max_count = 0
                    for sep in separators:
                        count = first_line.count(sep)
                        if count > max_count:
                            max_count = count
                            best_sep = sep
                    return enc, best_sep
            return enc, ','
        except UnicodeDecodeError:
            continue
    return 'utf-8', ','

def count_csv_rows(file_path):
    if not os.path.exists(file_path):
        return {"error": "File not found"}
    enc, sep = get_encoding_and_separator(file_path)
    count = 0
    try:
        with open(file_path, 'r', encoding=enc) as f:
            for _ in f:
                count += 1
        return {"status": "success", "count": max(0, count - 1)}
    except Exception as e:
        return {"error": str(e)}

def preview_csv(file_path):
    if not os.path.exists(file_path):
        return {"error": "File not found"}
    
    enc, sep = get_encoding_and_separator(file_path)
    try:
        with open(file_path, 'r', encoding=enc) as f:
            reader = csv.reader(f, delimiter=sep)
            headers = next(reader, [])
            sample = []
            for i, row in enumerate(reader):
                if i >= 5: # get 5 rows sample
                    break
                sample.append(dict(zip(headers, row)))
            
            # Estimate rows (very roughly by file size / avg line size)
            f.seek(0, os.SEEK_END)
            size = f.tell()
            est_rows = max(1, size // 150) # Rough estimate

        return {
            "type": "spreadsheet",
            "estimatedRows": est_rows,
            "columnsCount": len(headers),
            "separator": sep,
            "encoding": enc,
            "columns": headers,
            "sample": sample,
            "notice": "Tabela CSV identificada."
        }
    except Exception as e:
        return {"error": str(e)}

def extract_csv(file_path, filter_conditions, selected_columns, output_path, start_row=None, end_row=None):
    from sys import stdout
    enc, sep = get_encoding_and_separator(file_path)
    rows_extracted = 0
    try:
        # Get total rows for progress roughly
        total_rows = 1
        try:
            with open(file_path, 'r', encoding=enc) as f:
                total_rows = sum(1 for _ in f) - 1
            if total_rows < 1: total_rows = 1
        except:
             pass

        with open(file_path, 'r', encoding=enc) as infile, open(output_path, 'w', encoding='utf-8', newline='') as outfile:
            reader = csv.DictReader(infile, delimiter=sep)
            out_headers = selected_columns if selected_columns else reader.fieldnames
            writer = csv.DictWriter(outfile, fieldnames=out_headers)
            writer.writeheader()
            
            for i, row in enumerate(reader):
                if i % max(1, total_rows // 10) == 0:
                    print(json.dumps({"type": "progress", "percent": int((i / total_rows) * 100)}))
                    stdout.flush()

                if start_row is not None and i < start_row:
                    continue
                if end_row is not None and i > end_row:
                    break
                    
                match = True
                if filter_conditions:
                    for k, v in filter_conditions.items():
                        if str(row.get(k, '')).lower() != str(v).lower():
                            match = False
                            break
                if match:
                    out_row = {k: row.get(k) for k in out_headers} if selected_columns else row
                    writer.writerow(out_row)
                    rows_extracted += 1
                    
        res = {"status": "success", "rows_extracted": rows_extracted, "output_path": output_path}
        print(json.dumps({"type": "result", "data": res}))
        stdout.flush()
        return res
    except Exception as e:
        err = {"error": str(e)}
        print(json.dumps({"type": "result", "data": err}))
        stdout.flush()
        return err

def preview_xlsx(file_path):
    if not openpyxl:
        return {"error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        tabs = wb.sheetnames
        ws = wb[tabs[0]]
        
        headers = []
        sample = []
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(row)]
            elif i <= 5:
                row_dict = {}
                for j, h in enumerate(headers):
                    val = row[j] if j < len(row) else ""
                    row_dict[h] = str(val) if val is not None else ""
                sample.append(row_dict)
            else:
                break
                
        return {
            "type": "spreadsheet",
            "estimatedRows": ws.max_row if ws.max_row is not None else -1,
            "tabs": tabs,
            "columnsCount": len(headers),
            "columns": headers,
            "sample": sample,
            "notice": f"Planilha Excel identificada com {len(tabs)} abas."
        }
    except Exception as e:
        return {"error": str(e)}

def extract_xlsx(file_path, filter_conditions, selected_columns, sheet_name, output_path, output_format, start_row=None, end_row=None):
    from sys import stdout
    if not openpyxl:
        err = {"error": "openpyxl not installed"}
        print(json.dumps({"type": "result", "data": err}))
        stdout.flush()
        return err
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        total_rows = ws.max_row if ws.max_row else 1000

        # prepare csv writer if output format is csv
        is_csv = (output_format == 'csv')
        
        headers = []
        out_headers = []
        rows_extracted = 0
        
        if is_csv:
            outfile = open(output_path, 'w', encoding='utf-8', newline='')
            writer = None
        else:
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i % max(1, total_rows // 10) == 0:
                print(json.dumps({"type": "progress", "percent": int((i / total_rows) * 100)}))
                stdout.flush()

            if i == 0:
                headers = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(row)]
                out_headers = selected_columns if selected_columns else headers
                if is_csv:
                    writer = csv.DictWriter(outfile, fieldnames=out_headers)
                    writer.writeheader()
                else:
                    out_ws.append(out_headers)
                continue
                
            # offset for data 
            data_idx = i - 1
            if start_row is not None and data_idx < start_row:
                continue
            if end_row is not None and data_idx > end_row:
                break
                
            row_dict = {}
            for j, h in enumerate(headers):
                val = row[j] if j < len(row) else ""
                row_dict[h] = str(val) if val is not None else ""
                
            match = True
            if filter_conditions:
                for k, v in filter_conditions.items():
                    if str(row_dict.get(k, '')).lower() != str(v).lower():
                        match = False
                        break
                        
            if match:
                out_row_dict = {k: row_dict.get(k, '') for k in out_headers}
                if is_csv:
                    writer.writerow(out_row_dict)
                else:
                    out_ws.append([out_row_dict[k] for k in out_headers])
                rows_extracted += 1

        if is_csv:
            outfile.close()
        else:
            out_wb.save(output_path)
            
        res = {"status": "success", "rows_extracted": rows_extracted, "output_path": output_path}
        print(json.dumps({"type": "result", "data": res}))
        stdout.flush()
        return res
    except Exception as e:
        err = {"error": str(e)}
        print(json.dumps({"type": "result", "data": err}))
        stdout.flush()
        return err

if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            instructions = json.loads(sys.argv[1])
            action = instructions.get("action")
            file_path = instructions.get("file_path")
            
            if action == "preview_csv":
                res = preview_csv(file_path)
            elif action == "count_csv_rows":
                res = count_csv_rows(file_path)
            elif action == "extract_csv":
                out = instructions.get("output_path")
                filters = instructions.get("filters")
                cols = instructions.get("columns")
                start_row = instructions.get("startRow")
                end_row = instructions.get("endRow")
                res = extract_csv(file_path, filters, cols, out, start_row, end_row)
            elif action == "preview_xlsx":
                res = preview_xlsx(file_path)
            elif action == "extract_xlsx":
                out = instructions.get("output_path")
                filters = instructions.get("filters")
                cols = instructions.get("columns")
                sheet = instructions.get("sheetName")
                out_fmt = instructions.get("outputFormat", "xlsx")
                start_row = instructions.get("startRow")
                end_row = instructions.get("endRow")
                res = extract_xlsx(file_path, filters, cols, sheet, out, out_fmt, start_row, end_row)
            else:
                res = {"error": "Unknown action"}
                
            print(json.dumps(res))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
    else:
        print(json.dumps({"error": "No instructions provided"}))
